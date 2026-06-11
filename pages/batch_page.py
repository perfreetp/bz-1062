import os
import csv
from datetime import datetime

from PySide6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QFrame, QPushButton,
                               QLineEdit, QComboBox, QTableWidget, QTableWidgetItem,
                               QHeaderView, QCheckBox, QMessageBox, QFileDialog,
                               QLabel, QInputDialog, QAbstractItemView)
from PySide6.QtCore import Qt
from PySide6.QtGui import QColor

from database import PlantManager

try:
    import openpyxl
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False


class BatchPage(QWidget):
    def __init__(self):
        super().__init__()
        self.init_ui()
        self.load_data()

    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(12)

        toolbar = QFrame()
        toolbar.setStyleSheet('background: white; border-radius: 8px;')
        tb_layout = QHBoxLayout(toolbar)
        tb_layout.setContentsMargins(12, 10, 12, 10)
        tb_layout.setSpacing(10)

        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText('🔍 搜索植株...')
        self.search_input.setFixedWidth(220)
        self.search_input.returnPressed.connect(self.load_data)
        tb_layout.addWidget(self.search_input)

        self.status_combo = QComboBox()
        self.status_combo.addItem('全部状态')
        self.status_combo.addItems(['正常', '需关注', '病虫害', '枯死'])
        self.status_combo.setFixedWidth(120)
        self.status_combo.currentTextChanged.connect(self.load_data)
        tb_layout.addWidget(self.status_combo)

        self.area_combo = QComboBox()
        self.area_combo.addItem('全部区域')
        self.area_combo.setFixedWidth(140)
        tb_layout.addWidget(self.area_combo)

        btn_search = QPushButton('🔍 筛选')
        btn_search.clicked.connect(self.load_data)
        tb_layout.addWidget(btn_search)

        tb_layout.addStretch()

        btn_import = QPushButton('📥 导入Excel')
        btn_import.clicked.connect(self.import_data)
        tb_layout.addWidget(btn_import)

        btn_export = QPushButton('📤 导出Excel')
        btn_export.clicked.connect(self.export_data)
        tb_layout.addWidget(btn_export)

        btn_template = QPushButton('📄 下载模板')
        btn_template.clicked.connect(self.download_template)
        tb_layout.addWidget(btn_template)

        layout.addWidget(toolbar)

        batch_bar = QFrame()
        batch_bar.setStyleSheet('background: white; border-radius: 8px;')
        bb_layout = QHBoxLayout(batch_bar)
        bb_layout.setContentsMargins(12, 10, 12, 10)
        bb_layout.setSpacing(10)

        self.select_all_cb = QCheckBox('全选')
        self.select_all_cb.stateChanged.connect(self.toggle_select_all)
        bb_layout.addWidget(self.select_all_cb)

        self.selected_label = QLabel('已选择 0 项')
        self.selected_label.setStyleSheet('color: #606266;')
        bb_layout.addWidget(self.selected_label)

        bb_layout.addStretch()

        bb_layout.addWidget(QLabel('批量修改字段：'))
        self.field_combo = QComboBox()
        self.field_combo.addItems(['状态', '责任人', '所在区域', '品种', '规格'])
        self.field_combo.setFixedWidth(120)
        bb_layout.addWidget(self.field_combo)

        self.value_input = QLineEdit()
        self.value_input.setPlaceholderText('输入新值...')
        self.value_input.setFixedWidth(180)
        bb_layout.addWidget(self.value_input)

        btn_batch = QPushButton('⚡ 批量修改')
        btn_batch.setProperty('class', 'warning')
        btn_batch.clicked.connect(self.batch_update)
        bb_layout.addWidget(btn_batch)

        btn_delete = QPushButton('🗑️ 批量删除')
        btn_delete.setProperty('class', 'danger')
        btn_delete.clicked.connect(self.batch_delete)
        bb_layout.addWidget(btn_delete)

        layout.addWidget(batch_bar)

        self.table = QTableWidget()
        self.table.setEditTriggers(QAbstractItemView.DoubleClicked | QAbstractItemView.EditKeyPressed)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.NoSelection)
        self.table.itemChanged.connect(self.on_item_changed)

        headers = ['选择', 'ID', '名称', '品种', '规格', '数量', '面积(㎡)', '区域', '责任人', '状态', '种植日期', '备注']
        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels(headers)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.table.setColumnWidth(0, 50)
        self.table.setColumnWidth(1, 60)
        self.table.verticalHeader().setVisible(False)
        self.table.itemChanged.connect(self.update_selected_count)

        layout.addWidget(self.table, 1)

        hint = QLabel('💡 提示：双击单元格可直接编辑，勾选后可批量操作')
        hint.setStyleSheet('color: #909399; font-size: 12px;')
        layout.addWidget(hint)

    def load_data(self):
        self.table.blockSignals(True)
        keyword = self.search_input.text().strip() or None
        status = self.status_combo.currentText()
        if status == '全部状态':
            status = None
        area = self.area_combo.currentText()
        if area == '全部区域':
            area = None

        plants = PlantManager.get_all(keyword=keyword, status=status, area_name=area)
        self.current_data = plants

        self.table.setRowCount(0)
        status_colors = {
            '正常': QColor('#67c23a'),
            '需关注': QColor('#e6a23c'),
            '病虫害': QColor('#f56c6c'),
            '枯死': QColor('#909399'),
        }

        for p in plants:
            row = self.table.rowCount()
            self.table.insertRow(row)

            cb_item = QTableWidgetItem()
            cb_item.setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled)
            cb_item.setCheckState(Qt.Unchecked)
            cb_item.setData(Qt.UserRole, p['id'])
            self.table.setItem(row, 0, cb_item)

            self.table.setItem(row, 1, QTableWidgetItem(str(p['id'])))
            self.table.setItem(row, 2, QTableWidgetItem(p['name']))
            self.table.setItem(row, 3, QTableWidgetItem(p['species'] or ''))
            self.table.setItem(row, 4, QTableWidgetItem(p['spec'] or ''))
            self.table.setItem(row, 5, QTableWidgetItem(str(p['quantity'])))
            self.table.setItem(row, 6, QTableWidgetItem(str(p['area'])))
            self.table.setItem(row, 7, QTableWidgetItem(p['area_name'] or ''))
            self.table.setItem(row, 8, QTableWidgetItem(p['responsible'] or ''))

            status_item = QTableWidgetItem(p['status'] or '正常')
            color = status_colors.get(p['status'], QColor('#606266'))
            status_item.setForeground(color)
            self.table.setItem(row, 9, status_item)

            self.table.setItem(row, 10, QTableWidgetItem(p['plant_date'] or ''))
            self.table.setItem(row, 11, QTableWidgetItem(p['notes'] or ''))

        self.table.blockSignals(False)

        areas = PlantManager.get_areas()
        current_area = self.area_combo.currentText()
        self.area_combo.blockSignals(True)
        self.area_combo.clear()
        self.area_combo.addItem('全部区域')
        for a in areas:
            self.area_combo.addItem(a)
        if current_area:
            idx = self.area_combo.findText(current_area)
            if idx >= 0:
                self.area_combo.setCurrentIndex(idx)
        self.area_combo.blockSignals(False)

        self.update_selected_count()

    def toggle_select_all(self, state):
        self.table.blockSignals(True)
        check_state = Qt.Checked if state == Qt.Checked else Qt.Unchecked
        for row in range(self.table.rowCount()):
            item = self.table.item(row, 0)
            if item:
                item.setCheckState(check_state)
        self.table.blockSignals(False)
        self.update_selected_count()

    def get_selected_ids(self):
        ids = []
        for row in range(self.table.rowCount()):
            item = self.table.item(row, 0)
            if item and item.checkState() == Qt.Checked:
                ids.append(item.data(Qt.UserRole))
        return ids

    def update_selected_count(self):
        count = len(self.get_selected_ids())
        self.selected_label.setText(f'已选择 {count} 项')

    def on_item_changed(self, item):
        if item.column() == 0:
            return
        row = item.row()
        plant_id = self.table.item(row, 0).data(Qt.UserRole)
        field_map = {
            2: 'name', 3: 'species', 4: 'spec',
            5: 'quantity', 6: 'area', 7: 'area_name',
            8: 'responsible', 9: 'status', 10: 'plant_date', 11: 'notes'
        }
        field = field_map.get(item.column())
        if field:
            value = item.text()
            if field in ['quantity', 'area']:
                try:
                    value = float(value) if field == 'area' else int(value)
                except ValueError:
                    return
            PlantManager.update(plant_id, {field: value})

    def batch_update(self):
        ids = self.get_selected_ids()
        if not ids:
            QMessageBox.warning(self, '提示', '请先选择要修改的植株')
            return

        value = self.value_input.text().strip()
        if not value:
            QMessageBox.warning(self, '提示', '请输入新值')
            return

        field_map = {
            '状态': 'status',
            '责任人': 'responsible',
            '所在区域': 'area_name',
            '品种': 'species',
            '规格': 'spec',
        }
        field_label = self.field_combo.currentText()
        field = field_map.get(field_label)
        if not field:
            return

        reply = QMessageBox.question(self, '确认',
                                     f'确定要将选中的 {len(ids)} 项的{field_label}修改为 "{value}" 吗？')
        if reply == QMessageBox.Yes:
            PlantManager.batch_update(ids, field, value)
            QMessageBox.information(self, '成功', f'已批量修改 {len(ids)} 条记录')
            self.load_data()

    def batch_delete(self):
        ids = self.get_selected_ids()
        if not ids:
            QMessageBox.warning(self, '提示', '请先选择要删除的植株')
            return

        reply = QMessageBox.question(self, '确认',
                                     f'确定要删除选中的 {len(ids)} 株植物吗？\n（可在报表中心回收站恢复）')
        if reply == QMessageBox.Yes:
            PlantManager.batch_delete(ids)
            QMessageBox.information(self, '成功', f'已删除 {len(ids)} 条记录')
            self.load_data()

    def import_data(self):
        if not HAS_EXCEL:
            QMessageBox.warning(self, '提示', '未安装openpyxl库，仅支持CSV格式导入\n请运行: pip install openpyxl')
            file_filter = 'CSV文件 (*.csv)'
        else:
            file_filter = 'Excel文件 (*.xlsx *.xls);;CSV文件 (*.csv)'

        file_path, _ = QFileDialog.getOpenFileName(self, '选择导入文件', '', file_filter)
        if not file_path:
            return

        try:
            count = 0
            if file_path.endswith('.csv'):
                count = self._import_csv(file_path)
            else:
                count = self._import_excel(file_path)

            QMessageBox.information(self, '成功', f'成功导入 {count} 条记录')
            self.load_data()
        except Exception as e:
            QMessageBox.critical(self, '错误', f'导入失败：{str(e)}')

    def _import_csv(self, file_path):
        count = 0
        with open(file_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            for row in reader:
                data = {}
                for k, v in row.items():
                    k = k.strip()
                    if k in ['名称', 'name']:
                        data['name'] = v.strip()
                    elif k in ['品种', 'species']:
                        data['species'] = v.strip()
                    elif k in ['规格', 'spec']:
                        data['spec'] = v.strip()
                    elif k in ['数量', 'quantity']:
                        try:
                            data['quantity'] = int(v)
                        except:
                            data['quantity'] = 1
                    elif k in ['面积', 'area']:
                        try:
                            data['area'] = float(v)
                        except:
                            data['area'] = 0
                    elif k in ['区域', '所在区域', 'area_name']:
                        data['area_name'] = v.strip()
                    elif k in ['责任人', 'responsible']:
                        data['responsible'] = v.strip()
                    elif k in ['状态', 'status']:
                        data['status'] = v.strip()
                    elif k in ['种植日期', 'plant_date']:
                        data['plant_date'] = v.strip()
                    elif k in ['备注', 'notes']:
                        data['notes'] = v.strip()

                if 'name' in data and data['name']:
                    PlantManager.add(data)
                    count += 1
        return count

    def _import_excel(self, file_path):
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]

        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            data = {}
            for i, val in enumerate(row):
                if i >= len(headers):
                    break
                h = str(headers[i]).strip() if headers[i] else ''
                if h in ['名称', 'name']:
                    data['name'] = str(val).strip() if val else ''
                elif h in ['品种', 'species']:
                    data['species'] = str(val).strip() if val else ''
                elif h in ['规格', 'spec']:
                    data['spec'] = str(val).strip() if val else ''
                elif h in ['数量', 'quantity']:
                    try:
                        data['quantity'] = int(val)
                    except:
                        data['quantity'] = 1
                elif h in ['面积', 'area']:
                    try:
                        data['area'] = float(val)
                    except:
                        data['area'] = 0
                elif h in ['区域', '所在区域', 'area_name']:
                    data['area_name'] = str(val).strip() if val else ''
                elif h in ['责任人', 'responsible']:
                    data['responsible'] = str(val).strip() if val else ''
                elif h in ['状态', 'status']:
                    data['status'] = str(val).strip() if val else '正常'
                elif h in ['种植日期', 'plant_date']:
                    if hasattr(val, 'strftime'):
                        data['plant_date'] = val.strftime('%Y-%m-%d')
                    else:
                        data['plant_date'] = str(val).strip() if val else ''
                elif h in ['备注', 'notes']:
                    data['notes'] = str(val).strip() if val else ''

            if 'name' in data and data['name']:
                PlantManager.add(data)
                count += 1

        wb.close()
        return count

    def export_data(self):
        if not HAS_EXCEL:
            QMessageBox.warning(self, '提示', '未安装openpyxl库，导出为CSV格式')
            self._export_csv()
            return

        file_path, _ = QFileDialog.getSaveFileName(self, '导出为Excel',
                                                    f'绿植台账_{datetime.now().strftime("%Y%m%d")}.xlsx',
                                                    'Excel文件 (*.xlsx)')
        if not file_path:
            return

        try:
            self._export_excel(file_path)
            QMessageBox.information(self, '成功', f'导出成功，共 {len(self.current_data)} 条记录')
        except Exception as e:
            QMessageBox.critical(self, '错误', f'导出失败：{str(e)}')

    def _export_csv(self):
        file_path, _ = QFileDialog.getSaveFileName(self, '导出为CSV',
                                                    f'绿植台账_{datetime.now().strftime("%Y%m%d")}.csv',
                                                    'CSV文件 (*.csv)')
        if not file_path:
            return

        try:
            with open(file_path, 'w', encoding='utf-8-sig', newline='') as f:
                writer = csv.writer(f)
                writer.writerow(['名称', '品种', '规格', '数量', '面积(㎡)', '区域', '责任人', '状态', '种植日期', '备注'])
                for p in self.current_data:
                    writer.writerow([
                        p['name'], p.get('species', ''), p.get('spec', ''),
                        p.get('quantity', 1), p.get('area', 0),
                        p.get('area_name', ''), p.get('responsible', ''),
                        p.get('status', '正常'), p.get('plant_date', ''),
                        p.get('notes', '')
                    ])
            QMessageBox.information(self, '成功', f'导出成功，共 {len(self.current_data)} 条记录')
        except Exception as e:
            QMessageBox.critical(self, '错误', f'导出失败：{str(e)}')

    def _export_excel(self, file_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '绿植台账'

        headers = ['名称', '品种', '规格', '数量', '面积(㎡)', '区域', '责任人', '状态', '种植日期', '备注']
        ws.append(headers)

        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color='E8F4FD', end_color='E8F4FD', fill_type='solid')

        for p in self.current_data:
            ws.append([
                p['name'], p.get('species', ''), p.get('spec', ''),
                p.get('quantity', 1), p.get('area', 0),
                p.get('area_name', ''), p.get('responsible', ''),
                p.get('status', '正常'), p.get('plant_date', ''),
                p.get('notes', '')
            ])

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 4, 30)

        wb.save(file_path)
        wb.close()

    def download_template(self):
        file_path, _ = QFileDialog.getSaveFileName(self, '保存导入模板', '绿植台账导入模板.csv', 'CSV文件 (*.csv)')
        if not file_path:
            return

        try:
            with open(file_path, 'w', encoding='utf-8-sig', newline='') as f:
                writer = csv.writer(f)
                writer.writerow(['名称', '品种', '规格', '数量', '面积', '区域', '责任人', '状态', '种植日期', '备注'])
                writer.writerow(['示例：香樟树', '樟科', '胸径15cm', '1', '5.2', '东门广场', '张三', '正常', '2023-01-15', '示例数据'])
            QMessageBox.information(self, '成功', '模板下载成功')
        except Exception as e:
            QMessageBox.critical(self, '错误', f'下载失败：{str(e)}')

    def refresh(self):
        self.load_data()
